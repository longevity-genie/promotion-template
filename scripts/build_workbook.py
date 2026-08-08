#!/usr/bin/env python3
"""Build a spreadsheet view of the registry from the CSVs.

The CSVs are the source of truth, because they diff and merge in git. This script
produces a read-friendly .xlsx for people who would rather scan a spreadsheet.
The output is gitignored on purpose: edit the CSVs, regenerate this.

Usage:  python scripts/build_workbook.py [-o path/to/output.xlsx]
Requires openpyxl (pip install openpyxl).
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

# Relative to this file, so the repo works wherever it is cloned.
REPO = Path(__file__).resolve().parent.parent
REGISTRY = REPO / "registry"

# Sheet order matches the flow: what you write -> where it goes -> what happened.
SHEETS = [
    ("pillars", "pillars.csv"),
    ("derivatives", "derivatives.csv"),
    ("destinations", "destinations.csv"),
    ("shares", "shares.csv"),
    ("platform_rules", "platform_rules.csv"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F3350")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
TIER_FILLS = {
    "1": PatternFill("solid", fgColor="FDE7E9"),
    "2": PatternFill("solid", fgColor="FFF6E5"),
    "3": PatternFill("solid", fgColor="EAF7EE"),
}
INACTIVE_FILL = PatternFill("solid", fgColor="E8E8E8")
WIDE_COLUMNS = {"self_promo_rule", "notes", "key_rule_to_remember", "full_text",
                "text_sent", "thesis_one_line", "key_facts_to_include", "link_behaviour"}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("-o", "--output", default=str(REGISTRY / "registry_view.xlsx"))
    args = ap.parse_args()

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, filename in SHEETS:
        path = REGISTRY / filename
        if not path.exists():
            print(f"skipping {filename} (not found)")
            continue

        with path.open(newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        if not rows:
            continue

        header, body = rows[0], rows[1:]
        ws = wb.create_sheet(sheet_name)
        ws.append(header)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        idx = {name: i for i, name in enumerate(header)}
        for row in body:
            if not any(v.strip() for v in row):
                continue
            ws.append(row)
            r = ws.max_row
            # Tint by tier so priority is visible at a glance; grey out anything
            # not currently postable.
            fill = None
            if "status" in idx and row[idx["status"]] in {"on-hold", "comment-only", "retired"}:
                fill = INACTIVE_FILL
            elif "tier" in idx:
                fill = TIER_FILLS.get(row[idx["tier"]])
            for c in range(1, len(header) + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = BODY_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if fill:
                    cell.fill = fill

        for i, name in enumerate(header, start=1):
            width = 60 if name in WIDE_COLUMNS else max(12, min(28, len(name) + 8))
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"
        print(f"{sheet_name}: {ws.max_row - 1} rows")

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    # No formulas are written anywhere, so the file needs no recalculation pass.
    print(f"\nwrote {out}")
    print("The CSVs remain the source of truth - edit those, not this file.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
